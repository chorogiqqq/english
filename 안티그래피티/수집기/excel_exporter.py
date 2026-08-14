import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel(all_data, filepath):
    wb = openpyxl.Workbook()
    
    # 1. 통합 비교표 Sheet (Default sheet)
    ws_integrated = wb.active
    ws_integrated.title = '통합 비교표'
    
    # 2. 행정안전부 Sheet
    ws_mois = wb.create_sheet('행정안전부')
    
    # 3. 중소벤처기업부 Sheet
    ws_mss = wb.create_sheet('중소벤처기업부')
    
    # 4. 고용노동부 Sheet
    ws_moel = wb.create_sheet('고용노동부')
    
    # Prepare combined data sorted by registration date descending
    combined_list = []
    for ministry, items in all_data.items():
        combined_list.extend(items)
    combined_list.sort(key=lambda x: x.get('등록일', ''), reverse=True)

    sheet_mapping = [
        (ws_integrated, combined_list),
        (ws_mois, all_data.get('행정안전부', [])),
        (ws_mss, all_data.get('중소벤처기업부', [])),
        (ws_moel, all_data.get('고용노동부', []))
    ]

    # Style Definitions
    font_header = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid') # Dark Navy/Slate
    
    font_data = Font(name='맑은 고딕', size=10, color='0F172A')
    font_link = Font(name='맑은 고딕', size=10, color='2563EB', underline='single')
    
    fill_even = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_odd = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    headers = ['순번', '부처명', '공고번호', '등록일', '매칭키워드', '공고제목', '담당부서', '원문링크']

    for ws, data_list in sheet_mapping:
        ws.views.sheetView[0].showGridLines = True
        
        # Write Headers
        ws.row_dimensions[1].height = 28
        for col_idx, h_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
            
        # Write Data
        for row_idx, item in enumerate(data_list, 2):
            ws.row_dimensions[row_idx].height = 22
            row_fill = fill_even if row_idx % 2 == 0 else fill_odd
            
            link_url = item.get('원문링크', '')
            
            row_values = [
                row_idx - 1,                      # 순번
                item.get('부처명', ''),           # 부처명
                item.get('번호', ''),             # 공고번호
                item.get('등록일', ''),           # 등록일
                item.get('매칭키워드', ''),       # 매칭키워드
                item.get('제목', ''),             # 공고제목
                item.get('담당부서', ''),         # 담당부서
                "바로가기"                        # 원문링크
            ]
            
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                cell.border = thin_border
                
                if col_idx in [1, 2, 3, 4, 5]:
                    cell.font = font_data
                    cell.alignment = align_center
                elif col_idx in [6, 7]:
                    cell.font = font_data
                    cell.alignment = align_left
                elif col_idx == 8:
                    cell.font = font_link
                    cell.alignment = align_center
                    if link_url:
                        cell.hyperlink = link_url

        # Freeze Top Row
        ws.freeze_panes = 'A2'
        
        # Enable AutoFilter
        if len(data_list) > 0:
            ws.auto_filter.ref = f"A1:H{len(data_list)+1}"

        # Adjust Column Widths
        col_widths = {
            1: 8,   # 순번
            2: 15,  # 부처명
            3: 12,  # 공고번호
            4: 14,  # 등록일
            5: 18,  # 매칭키워드
            6: 55,  # 공고제목
            7: 20,  # 담당부서
            8: 12   # 원문링크
        }
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    wb.save(filepath)
    return filepath
